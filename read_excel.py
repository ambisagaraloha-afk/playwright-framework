from openpyxl import load_workbook

# Load the Excel file
file_path = r'd:\My Data\Automation-playwright\Playwright_Framework\test-cases_manual\DemoBlaze_TestCases.xlsx'
wb = load_workbook(file_path)
ws = wb.active

print(f'Sheet name: {ws.title}')
print()

# Get headers from first row
headers = []
for cell in ws[1]:
    headers.append(cell.value)

print('Available columns:')
print(headers)
print()

# Find TC-005 row
tc_005_row = None
for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=False), start=2):
    if row[0].value == 'TC-005':
        tc_005_row = row
        break

if tc_005_row:
    print('=' * 80)
    print('TEST CASE: TC-005 Details')
    print('=' * 80)
    
    for col_idx, header in enumerate(headers):
        value = tc_005_row[col_idx].value
        print(f'{header}: {value}')
    
    print('=' * 80)
else:
    print('TC-005 not found in the spreadsheet')
    print()
    print('Available Test Cases:')
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=1, values_only=True):
        print(f'  - {row[0]}')
